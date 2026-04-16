import io, datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.chart import BarChart, Reference


def create_batch_excel(batch_results, answer_key=None):

    wb = openpyxl.Workbook()

    hf = Font(bold=True, color="FFFFFF", size=11)
    hfill = PatternFill("solid", fgColor="4F46E5")
    cf = PatternFill("solid", fgColor="D1FAE5")
    wf = PatternFill("solid", fgColor="FEE2E2")
    bf = PatternFill("solid", fgColor="F1F5F9")
    border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))
    ct = Alignment(horizontal="center", vertical="center")
    lt = Alignment(horizontal="left", vertical="center")

  
    ws1 = wb.active
    ws1.title = "Ozet"

    headers1 = ["#", "Ad Soyad", "Ogrenci No", "Dogru", "Yanlis", "Bos", "Puan %", "Durum"]
    for col, h in enumerate(headers1, 1):
        c = ws1.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ct; c.border = border

    scores = []
    for i, br in enumerate(batch_results):
        row = i + 2
        si = br.get("student_info", {})
        comp = br.get("comparison", {})
        score = comp.get("score", 0)
        scores.append(score)

        ws1.cell(row=row, column=1, value=i+1).alignment = ct
        ws1.cell(row=row, column=2, value=si.get("name", "-")).alignment = lt
        ws1.cell(row=row, column=3, value=si.get("number", "-")).alignment = ct
        ws1.cell(row=row, column=4, value=comp.get("correct", 0)).alignment = ct
        ws1.cell(row=row, column=5, value=comp.get("wrong", 0)).alignment = ct
        ws1.cell(row=row, column=6, value=comp.get("blank", 0)).alignment = ct
        ws1.cell(row=row, column=7, value=score).alignment = ct

        if score >= 70:
            durum = "Basarili"
            fill = cf
        elif score >= 50:
            durum = "Orta"
            fill = PatternFill("solid", fgColor="FEF3C7")
        else:
            durum = "Basarisiz"
            fill = wf
        ws1.cell(row=row, column=8, value=durum).alignment = ct

        for col in range(1, len(headers1)+1):
            ws1.cell(row=row, column=col).border = border
        ws1.cell(row=row, column=7).fill = fill
        ws1.cell(row=row, column=8).fill = fill

    ws1.column_dimensions["A"].width = 6
    ws1.column_dimensions["B"].width = 24
    ws1.column_dimensions["C"].width = 16
    ws1.column_dimensions["D"].width = 10
    ws1.column_dimensions["E"].width = 10
    ws1.column_dimensions["F"].width = 8
    ws1.column_dimensions["G"].width = 10
    ws1.column_dimensions["H"].width = 12


    ws2 = wb.create_sheet("Detay")
    total_q = batch_results[0].get("total_questions", 0) if batch_results else 0

    dh = ["#", "Ad Soyad", "Ogrenci No"]
    for q in range(1, total_q + 1):
        dh.append(f"S{q}")
    dh.append("Puan")

    for col, h in enumerate(dh, 1):
        c = ws2.cell(row=1, column=col, value=h)
        c.font = hf; c.fill = hfill; c.alignment = ct; c.border = border


    if answer_key:
        row = 2
        ws2.cell(row=row, column=1, value="").alignment = ct
        ws2.cell(row=row, column=2, value="CEVAP ANAHTARI").font = Font(bold=True, color="4F46E5")
        ws2.cell(row=row, column=3, value="").alignment = ct
        for qi, key in enumerate(answer_key):
            ws2.cell(row=row, column=4+qi, value=key.upper()).alignment = ct
            ws2.cell(row=row, column=4+qi).font = Font(bold=True, color="4F46E5")
            ws2.cell(row=row, column=4+qi).border = border
        data_start = 3
    else:
        data_start = 2

    for i, br in enumerate(batch_results):
        row = data_start + i
        si = br.get("student_info", {})
        comp = br.get("comparison", {})
        answers = br.get("answers", [])

        ws2.cell(row=row, column=1, value=i+1).alignment = ct
        ws2.cell(row=row, column=2, value=si.get("name", "-")).alignment = lt
        ws2.cell(row=row, column=3, value=si.get("number", "-")).alignment = ct

        for qi, ans in enumerate(answers):
            cell = ws2.cell(row=row, column=4+qi, value=ans["answer"])
            cell.alignment = ct
            cell.border = border
            if comp and qi < len(comp.get("details", [])):
                d = comp["details"][qi]
                if d["status"] == "correct":
                    cell.fill = cf
                elif d["status"] == "wrong":
                    cell.fill = wf
                elif d["status"] == "blank":
                    cell.fill = bf

        ws2.cell(row=row, column=4+total_q, value=comp.get("score", 0)).alignment = ct
        for col in range(1, len(dh)+1):
            ws2.cell(row=row, column=col).border = border

    ws2.column_dimensions["A"].width = 6
    ws2.column_dimensions["B"].width = 24
    ws2.column_dimensions["C"].width = 16

 
    ws3 = wb.create_sheet("Istatistik")

   
    ws3.cell(row=1, column=1, value="SINIF ISTATISTIKLERI").font = Font(bold=True, size=14)

    if scores:
        import statistics
        stats_data = [
            ("Ogrenci Sayisi", len(scores)),
            ("Sinif Ortalamasi", f"{statistics.mean(scores):.1f}%"),
            ("En Yuksek Puan", f"{max(scores):.1f}%"),
            ("En Dusuk Puan", f"{min(scores):.1f}%"),
            ("Medyan", f"{statistics.median(scores):.1f}%"),
            ("Standart Sapma", f"{statistics.stdev(scores):.1f}%" if len(scores) > 1 else "-"),
            ("Basarili (>=70%)", f"{sum(1 for s in scores if s >= 70)} / {len(scores)}"),
            ("Basarisiz (<50%)", f"{sum(1 for s in scores if s < 50)} / {len(scores)}"),
        ]
        for j, (label, val) in enumerate(stats_data):
            r = 3 + j
            ws3.cell(row=r, column=1, value=label).font = Font(bold=True)
            ws3.cell(row=r, column=1).border = border
            ws3.cell(row=r, column=2, value=val).alignment = ct
            ws3.cell(row=r, column=2).border = border

  
    if answer_key and batch_results:
        sr = 14
        ws3.cell(row=sr, column=1, value="SORU BAZLI ANALIZ").font = Font(bold=True, size=14)

        qheaders = ["Soru", "Dogru Cevap", "Dogru %", "En Cok Verilen", "Bos %"]
        for col, h in enumerate(qheaders, 1):
            c = ws3.cell(row=sr+1, column=col, value=h)
            c.font = hf; c.fill = hfill; c.alignment = ct; c.border = border

        for qi in range(total_q):
            row = sr + 2 + qi
           
            given_answers = []
            correct_count = 0
            blank_count = 0
            for br in batch_results:
                comp = br.get("comparison", {})
                details = comp.get("details", [])
                if qi < len(details):
                    d = details[qi]
                    if d["status"] == "correct":
                        correct_count += 1
                    if d["given"] in ("BLANK", "BOS", ""):
                        blank_count += 1
                    else:
                        given_answers.append(d["given"])

            n = len(batch_results)
            correct_pct = round(correct_count / n * 100, 1) if n > 0 else 0
            blank_pct = round(blank_count / n * 100, 1) if n > 0 else 0

           
            if given_answers:
                from collections import Counter
                most_common = Counter(given_answers).most_common(1)[0][0]
            else:
                most_common = "-"

            expected = answer_key[qi].upper() if qi < len(answer_key) else "?"
            ws3.cell(row=row, column=1, value=qi+1).alignment = ct
            ws3.cell(row=row, column=2, value=expected).alignment = ct
            ws3.cell(row=row, column=3, value=correct_pct).alignment = ct
            ws3.cell(row=row, column=4, value=most_common).alignment = ct
            ws3.cell(row=row, column=5, value=blank_pct).alignment = ct

            for col in range(1, 6):
                ws3.cell(row=row, column=col).border = border
            if correct_pct < 30:
                ws3.cell(row=row, column=3).fill = wf
            elif correct_pct >= 70:
                ws3.cell(row=row, column=3).fill = cf

    ws3.column_dimensions["A"].width = 18
    ws3.column_dimensions["B"].width = 16
    ws3.column_dimensions["C"].width = 12
    ws3.column_dimensions["D"].width = 16
    ws3.column_dimensions["E"].width = 10

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf


def get_export_filename():
    return f"omr_toplu_sonuc_{datetime.date.today()}.xlsx"
